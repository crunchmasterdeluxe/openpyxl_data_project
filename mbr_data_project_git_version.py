import mysql.connector
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import plotly.express as px
import plotly.graph_objects as go
import datetime


allmonths = pd.DataFrame([
                        "5/2021",
                        "6/2021",
                        "7/2021",
                        "8/2021",
                        "9/2021",
                        "10/2021",
                        "11/2021",
                        "12/2021",
                        "1/2022",
                        "2/2022",
                        "3/2022",
                        "4/2022",
                        "5/2022"])
allmonths.columns = ['Month']

thick_border = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

month = '05'
sixmonths = '12'
thisyear = '2022'
lastyear = '2021'
sixmonthsbeginning = f'{lastyear}-12-01'
sixmonthsend = f'{thisyear}-02-01'
threemonthsbeginning = '2022-02-01'
threemonthsend = datetime.datetime.today().strftime('%Y-%m-%d')
monthname = 'May'
nextmonth = 'June'
today = datetime.datetime.today().strftime('%Y-%m-%d')

def makeInt(df,col):
    df[col] = df[col].fillna(0)
    df[col] = df[col].astype(int)

#PRODUCTION

db = mysql.connector.connect(

)
cursor = db.cursor()
cursor.execute(f"SELECT o.id office_id,o.name Office,COUNT(DISTINCT(a.employee_id)) active_reps,\
COUNT(a.customer_signoff) Signed,SUM(kw) signed_kw FROM lgcypower_panel_financial.accounts a \
LEFT JOIN lgcypower_panel_organization.offices o ON o.id = a.office_id \
LEFT JOIN lgcypower_panel_organization.employees e ON e.id = a.employee_id \
WHERE MONTH(a.customer_signoff) = '{month}' AND YEAR(a.customer_signoff) = '{thisyear}' \
AND a.dealer_id = 1 AND a.employee_id IS NOT NULL AND o.name IS NOT NULL \
GROUP BY o.name")
office_signed = cursor.fetchall()
df = pd.DataFrame(office_signed,columns=[i[0] for i in cursor.description])


for (i,j) in zip(offices['office_id'],offices['Office']):

    wb = Workbook()
    wb.remove(wb['Sheet'])
    sheet = wb.create_sheet("MBRs")
    active = wb["MBRs"]
    active.sheet_view.showGridLines = False

    img = Image('/Users/andy/Desktop/mountain3.png')
    img.width = 1900
    img.height = 800
    active.add_image(img, 'A1')

    df2=df.loc[df['office_id'] == i]
    df2=df2[['Active Reps','Signed','kWs Signed','Installed','kWs Installed']]

    sheet = wb.create_sheet("Production")
    active = wb["Production"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Production"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    active.merge_cells('I1:O1')
    active['I1'] = "* An active rep is one who has at least 1 signed deal in the month."

    active.column_dimensions['B'].width = 24
    active.column_dimensions['C'].width = 12
    active.column_dimensions['D'].width = 12
    active.column_dimensions['E'].width = 12
    active.column_dimensions['G'].width = 12

    active.cell(row=6, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=7).border = Border(bottom=Side(style='thin'))

    active['C6'].font = Font(bold=True)
    active['D6'].font = Font(bold=True)
    active['E6'].font = Font(bold=True)
    active['F6'].font = Font(bold=True)
    active['G6'].font = Font(bold=True)

    active['B7'] = f"{monthname} {thisyear}"

    rows = dataframe_to_rows(df2,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 3):
            active.cell(row=r_idx, column=c_idx, value=value)

    active.merge_cells('B7:B9')
    active.merge_cells('C7:C9')
    active.merge_cells('D7:D9')
    active.merge_cells('E7:E9')
    active.merge_cells('F7:F9')
    active.merge_cells('G7:G9')
    b7 = active['B7']
    b7.font = Font(size=12)
    b7.alignment = Alignment(vertical='center',horizontal='center')
    c7 = active['C7']
    c7.font = Font(size=18)
    c7.alignment = Alignment(vertical='center',horizontal='center')
    d6 = active['D6']
    d6.alignment = Alignment(horizontal='center')
    d7 = active['D7']
    d7.font = Font(size=18)
    d7.alignment = Alignment(vertical='center',horizontal='center')
    e7 = active['E7']
    e7.font = Font(size=18)
    e7.alignment = Alignment(vertical='center',horizontal='center')
    f7 = active['F7']
    f7.font = Font(size=18)
    f7.alignment = Alignment(vertical='center',horizontal='center')
    g7 = active['G7']
    g7.font = Font(size=18)
    g7.alignment = Alignment(vertical='center',horizontal='center')

    #Rep Age Segmentations
    active['A9'] = "Rep Age"
    active['A10'] = "Rookie"
    active['A11'] = "Vet"
    active['A12'] = "Dead"

    active['B10'] = "Reps Younger than 1 Quarter"
    active['B11'] = "Reps Older than 1 Quarter"
    active['B12'] = "Deactived"

    young_reps_2=young_reps.loc[young_reps['office_id'] == i]
    young_reps_2=young_reps_2[['Active Reps','Signed','kWs Signed','Installed','kWs Installed']]

    if len(young_reps_2) > 0:
        rows = dataframe_to_rows(young_reps_2,index=False,header=False)
        for r_idx, row in enumerate(rows, 10):
            for c_idx, value in enumerate(row, 3):
                active.cell(row=r_idx, column=c_idx, value=value)

    active.cell(row=7, column=2).border = Border(right=Side(style='thin'))
    active.cell(row=8, column=2).border = Border(right=Side(style='thin'))
    active.cell(row=9, column=2).border = Border(right=Side(style='thin'))
    active.cell(row=9, column=1).border = Border(bottom=Side(style='thin'))
    active.cell(row=9, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=9, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=9, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=9, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=9, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=9, column=7).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=1).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=10, column=7).border = Border(bottom=Side(style='thin'))
    active['C10'].alignment = Alignment(vertical='center',horizontal='center')
    active['D10'].alignment = Alignment(vertical='center',horizontal='center')
    active['E10'].alignment = Alignment(vertical='center',horizontal='center')
    active['F10'].alignment = Alignment(vertical='center',horizontal='center')
    active['G10'].alignment = Alignment(vertical='center',horizontal='center')

    old_reps_2=old_reps.loc[old_reps['office_id'] == i]
    old_reps_2=old_reps_2[['Active Reps','Signed','kWs Signed','Installed','kWs Installed']]

    if len(old_reps_2) > 0:
        rows = dataframe_to_rows(old_reps_2,index=False,header=False)
        for r_idx, row in enumerate(rows, 11):
            for c_idx, value in enumerate(row, 3):
                active.cell(row=r_idx, column=c_idx, value=value)

    active.cell(row=11, column=1).border = Border(bottom=Side(style='thin'))
    active.cell(row=11, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=11, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=11, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=11, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=11, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=11, column=7).border = Border(bottom=Side(style='thin'))
    active['C11'].alignment = Alignment(vertical='center',horizontal='center')
    active['D11'].alignment = Alignment(vertical='center',horizontal='center')
    active['E11'].alignment = Alignment(vertical='center',horizontal='center')
    active['F11'].alignment = Alignment(vertical='center',horizontal='center')
    active['G11'].alignment = Alignment(vertical='center',horizontal='center')

    dead_reps_2=dead_reps.loc[dead_reps['office_id'] == i]
    dead_reps_2=dead_reps_2[['Active Reps','Signed','kWs Signed','Installed','kWs Installed']]

    if len(dead_reps_2) > 0:
        rows = dataframe_to_rows(dead_reps_2,index=False,header=False)
        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 3):
                active.cell(row=r_idx, column=c_idx, value=value)

    active.cell(row=12, column=1).border = Border(bottom=Side(style='thin'))
    active.cell(row=12, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=12, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=12, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=12, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=12, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=12, column=7).border = Border(bottom=Side(style='thin'))
    active['C12'].alignment = Alignment(vertical='center',horizontal='center')
    active['D12'].alignment = Alignment(vertical='center',horizontal='center')
    active['E12'].alignment = Alignment(vertical='center',horizontal='center')
    active['F12'].alignment = Alignment(vertical='center',horizontal='center')
    active['G12'].alignment = Alignment(vertical='center',horizontal='center')

    #Historical Performance
    d=history.loc[history['office_id'] == i]
    dates=[]
    for (aa,bb) in zip(d['y'],d['m']):
        dates.append(f"{bb}/{aa}")
    d['Month']=dates

    d=pd.merge(allmonths,d,
              how='left',
              on='Month',
              indicator=False)
    d.fillna(0,inplace=True)
    makeInt(d,'Active Reps')
    makeInt(d,'Signed')
    makeInt(d,'Installed')

    fig = go.Figure()

    # Add traces
    fig.add_trace(go.Scatter(x=d['Month'], y=d['Signed'],
                        mode='lines+markers+text',
                        name='Signed',
                        marker=dict(color='#0DABE3'),
                        textfont=dict(color='#0DABE3'),
                        text=d['Signed'],
                        textposition='top center'))
    fig.add_trace(go.Scatter(x=d['Month'], y=d['Installed'],
                        mode='lines+markers+text',
                        name='Installed',
                        marker=dict(color="#1167B1"),
                        textfont=dict(color="#1167B1"),
                        text=d['Installed'],
                        textposition='top center'))
    fig.add_trace(go.Scatter(x=d['Month'], y=d['Active Reps'],
                        mode='lines+markers+text',
                        name='Active Reps',
                        marker=dict(color="#03254C"),
                        textfont=dict(color="#03254C"),
                        text=d['Active Reps'],
                        textposition='top center'))

    fig.update_layout(title='1 Year Overview',margin=dict(l=10, r=10, t=100, b=10),
                      legend=dict(orientation="h",
                                yanchor="bottom",
                                y=1.01,
                                xanchor="left",
                                x=0))
    fig.update_xaxes(
        dtick="M1",
        tickformat="%b\n%Y")
    fig.update_layout(legend= {'itemsizing': 'constant'})

    fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_overview.png")

    img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_overview.png")
    img.width = 850
    img.height = 600
    #active.merge_cells('I14:O30')
    active.add_image(img,'B15')

    #PRA and PARA
    df25=df24.loc[df24['office_id'] == i]
    count1=[]
    for uuu in df25['Signed']:
        if uuu > 0:
            count1.append(1)

    if len(df25) > 0:
        active['I6'] = f"Per Rep Avg (PRA)"
        active['I7'] = float(round((sum(df25['Signed'])/len(df25)),1))
        active['I7'].number_format = '0.0'
        active['J6'] = f"Per Active Rep Avg (PARA)"
        if len(count1) > 0:
            active['J7'] = float(round((sum(df25['Signed'])/len(count1)),1))
        else:
            active['J7'] = 0.0
        active['J7'].number_format = '0.0'
        active.cell(row=6, column=9).border = Border(bottom=Side(style='thin'))
        active.cell(row=6, column=10).border = Border(bottom=Side(style='thin'))
        active['I6'].font = Font(bold=True)
        active['J6'].font = Font(bold=True)
        active.column_dimensions['I'].width = 16
        active.column_dimensions['J'].width = 20
        active['I7'].font = Font(size=18)
        active['J7'].font = Font(size=18)
        active['I6'].alignment = Alignment(horizontal='center',vertical='center')
        active['I7'].alignment = Alignment(horizontal='center',vertical='center')
        active['J6'].alignment = Alignment(horizontal='center',vertical='center')
        active['J7'].alignment = Alignment(horizontal='center',vertical='center')
        active.merge_cells('I7:I9')
        active.merge_cells('J7:J9')


    #Scheduled for Install
    active.merge_cells('C33:F33')
    active['C33'] = f"{nextmonth} Scheduled Installs"
    active.cell(row=33, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=33, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=33, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=33, column=6).border = Border(bottom=Side(style='thin'))
    df39=df38.loc[df38['office_id'] == i]
    df39 = df39[['scheduled', 'kW']]
    df39 = df39.rename(index=str, columns={'scheduled':'Scheduled'})
    rows = dataframe_to_rows(df39,index=False)
    for r_idx, row in enumerate(rows, 34):
        for c_idx, value in enumerate(row, 4):
            active.cell(row=r_idx, column=c_idx, value=value)

    active.cell(row=34, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=34, column=5).border = Border(bottom=Side(style='thin'))
    active['C34'].font = Font(bold=True)
    active['D34'].font = Font(bold=True)

    active['C33'].alignment = Alignment(horizontal='center')
    active['D34'].alignment = Alignment(horizontal='center')
    active['D35'].alignment = Alignment(horizontal='center')
    active['E34'].alignment = Alignment(horizontal='center')
    active['E35'].alignment = Alignment(horizontal='center')


    ####################################################################################
    #####################################Partner Split##################################
    ####################################################################################
    sheet = wb.create_sheet("Install Partners")
    active = wb["Install Partners"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Install Partner Report"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    df7=partner2.loc[partner2['office_id'] == i]
    df7=df7[['Install Partner', 'Approved', 'Split','COBF','6mo Avg Cycle Time (Days)']]

    #find install partner split per office
    totalPartner = sum(df7['Approved'])
    partnerPerc=[]
    for jj in df7['Approved']:
        k=float(jj/totalPartner)
        partnerPerc.append(k)

    df7['Split'] = partnerPerc

    fig = px.pie(df7, values='Approved', names='Install Partner', color='Install Partner',
             color_discrete_sequence=["#0DABE3","#03254C", "#1167B1", "blue", "#187BCD", "#D0EFFF"],hole=.3)

    fig.update_layout(title='Partner Approval Split',margin=dict(l=10, r=10, t=100, b=10),
                      legend=dict(orientation="h",
                                yanchor="bottom",
                                y=1.01,
                                xanchor="left",
                                x=0),
                     font=dict(
                                size=18
                            ))

    fig.update_layout(uniformtext_minsize=20, uniformtext_mode='hide')

    fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_pie.png")

    img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_pie.png")
    img.width = 400
    img.height = 250
    #     active.merge_cells('I3:Q43')
    active.add_image(img,'J2')

    rows = dataframe_to_rows(df7,index=False)
    for r_idx, row in enumerate(rows, 7):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    active.cell(row=7, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=7).border = Border(bottom=Side(style='thin'))
    active['B7'].font = Font(bold=True)
    active['C7'].font = Font(bold=True)
    active['D7'].font = Font(bold=True)
    active['E7'].font = Font(bold=True)
    active['F7'].font = Font(bold=True)
    active['G7'].font = Font(bold=True)


    active.merge_cells('F7:G7')
    active.merge_cells('F8:G8')
    active.merge_cells('F9:G9')
    active.merge_cells('F10:G10')
    active.merge_cells('F11:G11')
    active.merge_cells('F12:G12')
    active.merge_cells('F13:G13')
    #     active['B25'].number_format = '0%'
    active['D8'].number_format = '0%'
    active['E8'].number_format = '0%'
    active['D9'].number_format = '0%'
    active['E9'].number_format = '0%'
    active['D10'].number_format = '0%'
    active['E10'].number_format = '0%'
    active['D11'].number_format = '0%'
    active['E11'].number_format = '0%'
    active['D12'].number_format = '0%'
    active['E12'].number_format = '0%'
    active['D13'].number_format = '0%'
    active['E13'].number_format = '0%'

    active['B7'].alignment = Alignment(horizontal='center')
    active['B8'].alignment = Alignment(horizontal='center')
    active['B9'].alignment = Alignment(horizontal='center')
    active['B10'].alignment = Alignment(horizontal='center')
    active['B11'].alignment = Alignment(horizontal='center')
    active['B12'].alignment = Alignment(horizontal='center')
    active['B13'].alignment = Alignment(horizontal='center')
    active['C7'].alignment = Alignment(horizontal='center')
    active['C8'].alignment = Alignment(horizontal='center')
    active['C9'].alignment = Alignment(horizontal='center')
    active['C10'].alignment = Alignment(horizontal='center')
    active['C11'].alignment = Alignment(horizontal='center')
    active['C12'].alignment = Alignment(horizontal='center')
    active['C13'].alignment = Alignment(horizontal='center')
    active['D7'].alignment = Alignment(horizontal='center')
    active['D8'].alignment = Alignment(horizontal='center')
    active['D9'].alignment = Alignment(horizontal='center')
    active['D10'].alignment = Alignment(horizontal='center')
    active['D11'].alignment = Alignment(horizontal='center')
    active['D12'].alignment = Alignment(horizontal='center')
    active['D13'].alignment = Alignment(horizontal='center')
    active['E7'].alignment = Alignment(horizontal='center')
    active['E8'].alignment = Alignment(horizontal='center')
    active['E9'].alignment = Alignment(horizontal='center')
    active['E10'].alignment = Alignment(horizontal='center')
    active['E11'].alignment = Alignment(horizontal='center')
    active['E12'].alignment = Alignment(horizontal='center')
    active['E13'].alignment = Alignment(horizontal='center')
    active['F7'].alignment = Alignment(horizontal='center')
    active['F8'].alignment = Alignment(horizontal='center')
    active['F9'].alignment = Alignment(horizontal='center')
    active['F10'].alignment = Alignment(horizontal='center')
    active['F11'].alignment = Alignment(horizontal='center')
    active['F12'].alignment = Alignment(horizontal='center')
    active['F13'].alignment = Alignment(horizontal='center')

    active.column_dimensions['B'].width = 12
    active.column_dimensions['F'].width = 18

    office_installer_scat = installer_scat.loc[installer_scat['office_id'] == i]
    if len(office_installer_scat) > 0:
        fig = px.scatter(office_installer_scat,x="Date", y="Days Signed to Install",color='Installer',trendline="lowess")
        #                 color_discrete_sequence=["#0DABE3","#03254C", "#1167B1", "blue", "#187BCD", "#D0EFFF"])
        tr_line=[]
        for  k, trace  in enumerate(fig.data):
            if trace.mode is not None and trace.mode == 'lines':
                tr_line.append(k)

        for id in tr_line:
            fig.data[id].update(line_width=4)

        fig.update_traces(marker=dict(size=3))
        fig.update_xaxes(
            dtick="M1",
            tickformat="%b\n%Y")

        fig.update_layout(title='Install Efficiency',margin=dict(l=10, r=10, t=100, b=10),
                          legend=dict(orientation="h",
                                    yanchor="bottom",
                                    y=1.01,
                                    xanchor="left",
                                    x=0))
        fig.update_layout(legend= {'itemsizing': 'constant'})


        fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_installer_scat.png")

        img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_installer_scat.png")
        img.width = 600
        img.height = 400
        #active.merge_cells('I14:O30')
        active.add_image(img,'B16')

    office_installer_prog = installer_prog.loc[installer_prog['office_id'] == i]
    if len(office_installer_prog) > 0:
        installers = office_installer_prog.drop_duplicates(['Installer'],keep='first')
        counter=[]
        for ii in installers['Installer']:
            num = 1
            for jj in office_installer_prog['Installer']:
                if jj == ii:
                    counter.append(num)
                    num += 1
        office_installer_prog['Installed'] = counter

        fig = px.line(office_installer_prog, x="Date", y="Installed", color='Installer')
        fig.update_layout(title=f'Install Tracker {thisyear}',margin=dict(l=10, r=10, t=100, b=10),
                          legend=dict(orientation="h",
                                    yanchor="bottom",
                                    y=1.01,
                                    xanchor="left",
                                    x=0))

        fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_installer_prog.png")

        img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_installer_prog.png")
        img.width = 600
        img.height = 400
        #active.merge_cells('I14:O30')
        active.add_image(img,'K16')



    ####################################################################################
    #####################################Pipeline#######################################
    ####################################################################################
    sheet = wb.create_sheet("Pipeline")
    active = wb["Pipeline"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Pipeline"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    df9=df8.loc[df8['office_id'] == i]
    df9=df9[['Signed to Approved','Signed to Approved','Approved to M1','Approved to M1','M1 to M2','M1 to M2']]

    df9_days = df8.loc[df8['office_id'] == i]
    df9_days = df9_days[['Days to Approval','Days to Approval','Days to M1','Days to M1','Days to M2','Days to M2']]

    df9_3mo=df8_3mo.loc[df8_3mo['office_id'] == i]
    df9_3mo=df9_3mo[['Signed to Approved','Signed to Approved','Approved to M1','Approved to M1','M1 to M2','M1 to M2']]

    active.merge_cells('B7:G7')
    active['B7'] = "Pull-Through - Cohort"
    active['B7'].alignment = Alignment(vertical='center',horizontal='center')

    if len(df9) > 0:
        df9['Signed to Approved'] = df9['Signed to Approved'].astype(float)
        df9['Approved to M1'] = df9['Approved to M1'].astype(float)
        df9['M1 to M2'] = df9['M1 to M2'].astype(float)

        rows = dataframe_to_rows(df9,index=False)
        for r_idx, row in enumerate(rows, 8):
            for c_idx, value in enumerate(row, 2):
                active.cell(row=r_idx, column=c_idx, value=value)

        rows = dataframe_to_rows(df9_days,index=False,header=False)
        for r_idx, row in enumerate(rows, 12):
            for c_idx, value in enumerate(row, 2):
                active.cell(row=r_idx, column=c_idx, value=value)


    rows = dataframe_to_rows(df10,index=False)
    for r_idx, row in enumerate(rows, 14):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    if len(df9_3mo) > 0:
        rows = dataframe_to_rows(df9_3mo,index=False)
        for r_idx, row in enumerate(rows, 17):
            for c_idx, value in enumerate(row, 2):
                active.cell(row=r_idx, column=c_idx, value=value)

    active.merge_cells('B8:C8')
    active.merge_cells('D8:E8')
    active.merge_cells('F8:G8')
    active.merge_cells('A9:A11')
    active['A9'] = "6mo-3mo"
    active['A9'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B9:C11')
    active.merge_cells('D9:E11')
    active.merge_cells('F9:G11')
    active['B9'].font = Font(size=16)
    active['B8'].font = Font(bold=True)
    active['B9'].number_format = '0%'
    active['B8'].alignment = Alignment(vertical='center',horizontal='center')
    active['B9'].alignment = Alignment(vertical='center',horizontal='center')
    active['D9'].font = Font(size=16)
    active['D8'].font = Font(bold=True)
    active['D9'].number_format = '0%'
    active['D8'].alignment = Alignment(vertical='center',horizontal='center')
    active['D9'].alignment = Alignment(vertical='center',horizontal='center')
    active['F9'].font = Font(size=16)
    active['F8'].font = Font(bold=True)
    active['F9'].number_format = '0%'
    active['F8'].alignment = Alignment(vertical='center',horizontal='center')

    active['F9'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('A12:A14')
    active['A12'] = "Avg Days"
    active['A12'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B12:C14')
    active['B12'].font = Font(size=16)
    active['B12'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('D12:E14')
    active['D12'].font = Font(size=16)
    active['D12'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('F12:G14')
    active['F12'].font = Font(size=16)
    active['F12'].alignment = Alignment(vertical='center',horizontal='center')

    active.merge_cells('A15:A17')
    active['A15'] = "LGCY Avg"
    active['A15'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B15:C17')
    active['B15'].font = Font(size=16)
    active['B15'].number_format = '0%'
    active['B15'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('D15:E17')
    active['D15'].font = Font(size=16)
    active['D15'].number_format = '0%'
    active['D15'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('F15:G17')
    active['F15'].font = Font(size=16)
    active['F15'].number_format = '0%'
    active['F15'].alignment = Alignment(vertical='center',horizontal='center')

    active.merge_cells('A18:A20')
    active['A18'] = "3mo-Now"
    active['A18'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B18:C20')
    active['B18'].font = Font(size=16)
    active['B18'].number_format = '0%'
    active['B18'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('D18:E20')
    active['D18'].font = Font(size=16)
    active['D18'].number_format = '0%'
    active['D18'].alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('F18:G20')
    active['F18'].number_format = '0%'
    active['F18'].font = Font(size=16)
    active['F18'].alignment = Alignment(vertical='center',horizontal='center')

    active.cell(row=7, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=7, column=7).border = Border(bottom=Side(style='thin'))
    active.cell(row=8, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=8, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=8, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=8, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=8, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=8, column=7).border = Border(bottom=Side(style='thin'))

    #         cells=list(range(32, (33 + newreps)))

    rep_pipe=pipe.loc[pipe['office_id'] == i]
    rep_pipe=rep_pipe.rename(columns={"Signed_Not_Approved":"Signed Not Approved","Approved_Not_M1":"Approved Not M1",
                             "M1_Not_M2":"M1 Not M2"})
    rep_pipe=rep_pipe[['Rep','Signed Not Approved','Approved Not M1','M1 Not M2','Total']]
    rep_pipe=rep_pipe.sort_values(by=['Total'], ascending=False)

    active.merge_cells('I5:M5')
    active['I5'] = "Un-Installed, Uncancelled - Last 3 Months"
    active['I5'].alignment = Alignment(horizontal='center')
    active['I6'].font = Font(bold=True)
    active['J6'].font = Font(bold=True)
    active['K6'].font = Font(bold=True)
    active['L6'].font = Font(bold=True)
    active['M6'].font = Font(bold=True)
    active.cell(row=5, column=9).border = Border(bottom=Side(style='thin'))
    active.cell(row=5, column=10).border = Border(bottom=Side(style='thin'))
    active.cell(row=5, column=11).border = Border(bottom=Side(style='thin'))
    active.cell(row=5, column=12).border = Border(bottom=Side(style='thin'))
    active.cell(row=5, column=13).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=9).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=10).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=11).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=12).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=13).border = Border(bottom=Side(style='thin'))
    affectedcells2 = list(range(6, 75))
    for rownum in affectedcells2:
        active[f'J{rownum}'].alignment = Alignment(horizontal='center')
        active[f'K{rownum}'].alignment = Alignment(horizontal='center')
        active[f'L{rownum}'].alignment = Alignment(horizontal='center')
        active[f'M{rownum}'].alignment = Alignment(horizontal='center')

    active.column_dimensions['I'].width = 14
    active.column_dimensions['J'].width = 20
    active.column_dimensions['K'].width = 20
    active.column_dimensions['L'].width = 16
    active.column_dimensions['M'].width = 16

    rows = dataframe_to_rows(rep_pipe,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 9):
            active.cell(row=r_idx, column=c_idx, value=value)

    ##########################################################################################
    ##########################################################################################
    #######################################New Reps###########################################
    ##########################################################################################
    ##########################################################################################

    sheet = wb.create_sheet("Recruiting")
    active = wb["Recruiting"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Recruiting"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    #Header
    active.merge_cells('I1:O1')
    active['I1'] = "* A new rep is one who had their first signed deal in the month."

    df15=df14.loc[df14['office_id'] == i]
    df17=df16.loc[df16['office_id'] == i]
    newreps=len(df17)
    if newreps > 0:
        active.merge_cells('I2:I4')
        active['I2'].alignment = Alignment(vertical='center',horizontal='center')
        active['I2'] = "New Reps"

        active.merge_cells('J2:J4')
        active['J2'].alignment = Alignment(vertical='center',horizontal='center')
        active['J2'].font = Font(size=18)
        active['J2'] = int(len(df17))
        active['J2'].number_format = '0'
    else:
        active.merge_cells('I3:K3')
        active['I3'] = "No New Rep Activity"

    #New Rep Activity
    df19=df18.loc[df18['office_id'] == i]
    df19=df19[['rep_name','Recruiter','Invited','welcome_complete','Signed']]
    df19=df19.rename(columns={"rep_name":"Rep","welcome_complete":"Onboarded"})

    df19=df19.sort_values(by=['Invited','Onboarded'],ascending=True)
    #df19=df19.sort_values(by=['Signed'],ascending=False)

    rows = dataframe_to_rows(df19,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    active['B6'].font = Font(bold=True)
    active['C6'].font = Font(bold=True)
    active['D6'].font = Font(bold=True)
    active['E6'].font = Font(bold=True)
    active['F6'].font = Font(bold=True)
    active.cell(row=6, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=6).border = Border(bottom=Side(style='thin'))
    active.column_dimensions['B'].width = 17
    active.column_dimensions['C'].width = 17
    active.column_dimensions['D'].width = 12
    active.column_dimensions['E'].width = 12
    active.column_dimensions['F'].width = 12
    affectedcells = list(range(6, 50))
    for rownum in affectedcells:
        active[f'D{rownum}'].alignment = Alignment(horizontal='center')
        active[f'E{rownum}'].alignment = Alignment(horizontal='center')
        active[f'F{rownum}'].alignment = Alignment(horizontal='center')

    d2=df22.loc[df22['office_id'] == i]
    dates=[]
    for (aaaa,bbbb) in zip(d2['y'],d2['m']):
        if bbbb == '01':
            bbbb = '1'
        if bbbb == '02':
            bbbb = '2'
        if bbbb == '03':
            bbbb = '3'
        if bbbb == '04':
            bbbb = '5'
        if bbbb == '05':
            bbbb = '5'
        if bbbb == '06':
            bbbb = '6'
        if bbbb == '07':
            bbbb = '7'
        if bbbb == '08':
            bbbb = '8'
        if bbbb == '09':
            bbbb = '9'
        dates.append(f"{bbbb}/{aaaa}")
    d2['Month']=dates

    allmonths = pd.DataFrame([
                        "5/2021",
                        "6/2021",
                        "7/2021",
                        "8/2021",
                        "9/2021",
                        "10/2021",
                        "11/2021",
                        "12/2021",
                        "1/2022",
                        "2/2022",
                        "3/2022",
                        "4/2022",
                        "5/2022"])
    allmonths.columns = ['Month']


    d2=pd.merge(allmonths,d2,
              how='left',
              on='Month',
              indicator=False)
    d2.fillna(0,inplace=True)

    d2['office_id'].fillna(i,inplace=True)
    d2.fillna(0,inplace=True)
    makeInt(d2,'office_id')

    d2=pd.merge(d2,m_reps,
           how='left',
           on=['office_id','Month'],
           indicator=False)
    makeInt(d2,'Invited')
    makeInt(d2,'Onboarded')
    d2['new_reps'].fillna(0,inplace=True)
    makeInt(d2,'new_reps')


    fig = go.Figure()

    # Add traces
    fig.add_trace(go.Scatter(x=d2['Month'], y=d2['Invited'],
                        mode='lines+markers+text',
                        name='Invited',
                        marker=dict(color='#0DABE3'),
                        textfont=dict(color='#0DABE3'),
                        text=d2['Invited'],
                        textposition='top center'))
    fig.add_trace(go.Scatter(x=d2['Month'], y=d2['Onboarded'],
                        mode='lines+markers+text',
                        name='Onboarded',
                        marker=dict(color="#1167B1"),
                        textfont=dict(color='#1167B1'),
                        text=d2['Onboarded'],
                        textposition='top center'))
    fig.add_trace(go.Scatter(x=d2['Month'], y=d2['new_reps'],
                        mode='lines+markers+text',
                        name='New Reps with Sale',
                        marker=dict(color="blue"),
                        textfont=dict(color='blue'),
                        text=d2['new_reps'],
                        textposition='top center'))


    fig.update_layout(title='1 Year Overview',margin=dict(l=10, r=10, t=100, b=10),
                      legend=dict(orientation="h",
                                yanchor="bottom",
                                y=1.01,
                                xanchor="left",
                                x=0))
    fig.update_xaxes(
        dtick="M1",
        tickformat="%b\n%Y")

    fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_recruiting_overview.png")

    img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_recruiting_overview.png")
    img.width = 700
    img.height = 550
    active.merge_cells('H6:P30')
    active.add_image(img,'H6')

    recruiters2 = recruiters.loc[recruiters['office_id'] == i]
    recruiters2 = recruiters2[['Recruiter','Invited','Onboarded','Reps with Signed Deal']]

    rows = dataframe_to_rows(recruiters2,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 19):
            active.cell(row=r_idx, column=c_idx, value=value)

    active['S6'].font = Font(bold=True)
    active['T6'].font = Font(bold=True)
    active['U6'].font = Font(bold=True)
    active['V6'].font = Font(bold=True)
    active.cell(row=6, column=19).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=20).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=21).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=22).border = Border(bottom=Side(style='thin'))
    active.column_dimensions['S'].width = 17
    active.column_dimensions['T'].width = 14
    active.column_dimensions['U'].width = 14
    active.column_dimensions['V'].width = 20
    affectedcells = list(range(6, 50))
    for rownum in affectedcells:
        active[f'T{rownum}'].alignment = Alignment(horizontal='center')
        active[f'U{rownum}'].alignment = Alignment(horizontal='center')
        active[f'V{rownum}'].alignment = Alignment(horizontal='center')



    ##########################################################################################
    ##########################################################################################
    #######################################Rep Activity#######################################
    ##########################################################################################
    ##########################################################################################
    sheet = wb.create_sheet("Rep Activity")
    active = wb["Rep Activity"]
    active.sheet_view.showGridLines = False

    active.column_dimensions['B'].width = 20
    active.column_dimensions['C'].width = 18
    active.column_dimensions['D'].width = 18
    active.column_dimensions['E'].width = 10
    active.column_dimensions['F'].width = 16
    active.column_dimensions['G'].width = 10
    active.column_dimensions['H'].width = 16
    active.column_dimensions['I'].width = 14
    active.column_dimensions['J'].width = 25
    active.column_dimensions['K'].width = 20
    active.column_dimensions['L'].width = 16
    active.column_dimensions['M'].width = 16
    active.column_dimensions['N'].width = 16
    active.column_dimensions['O'].width = 16
    affectedcells1 = list(range(6, 125))
    for rownum in affectedcells1:
        active[f'C{rownum}'].alignment = Alignment(horizontal='center')
        active[f'D{rownum}'].alignment = Alignment(horizontal='center')
        active[f'E{rownum}'].alignment = Alignment(horizontal='center')
        active[f'F{rownum}'].alignment = Alignment(horizontal='center')
        active[f'G{rownum}'].alignment = Alignment(horizontal='center')
        active[f'H{rownum}'].alignment = Alignment(horizontal='center')
        active[f'I{rownum}'].alignment = Alignment(horizontal='center')
        active[f'J{rownum}'].alignment = Alignment(horizontal='center')
        active[f'K{rownum}'].alignment = Alignment(horizontal='center')
        active[f'L{rownum}'].alignment = Alignment(horizontal='center')
        active[f'M{rownum}'].alignment = Alignment(horizontal='center')
        active[f'N{rownum}'].alignment = Alignment(horizontal='center')
        active[f'O{rownum}'].alignment = Alignment(horizontal='center')

    active['B6'].font = Font(bold=True)
    active['C6'].font = Font(bold=True)
    active['D6'].font = Font(bold=True)
    active['E6'].font = Font(bold=True)
    active['F6'].font = Font(bold=True)
    active['G6'].font = Font(bold=True)
    active['H6'].font = Font(bold=True)
    active['I6'].font = Font(bold=True)
    active['J6'].font = Font(bold=True)
    active['K6'].font = Font(bold=True)
    active['L6'].font = Font(bold=True)
    active['M6'].font = Font(bold=True)
    active['N6'].font = Font(bold=True)
    active['O6'].font = Font(bold=True)
    active.cell(row=6, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=6).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=7).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=8).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=9).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=10).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=11).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=12).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=13).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=14).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=15).border = Border(bottom=Side(style='thin'))

    active['B1'] = f"{j} Rep Activity"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    df25=df24.loc[df24['office_id'] == i]
    df25=df25[['Rep Name','Days with LGCY','Days to First Sale','Signed',
       'Daily Activity%', 'Opptys', 'Oppty Activity%', 'Interactions',
       'Daily Interactivity','Invited','Onboarded','Rep Rank',f'{nextmonth} # To Beat','All Time Best']]
    df25['Interactions'] = df25['Interactions'].astype(int)

    active.conditional_formatting.add('E7:E125',
                ColorScaleRule(start_type='percentile', start_value=10, start_color='FFFFFF',
                            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                            end_type='percentile', end_value=95, end_color='AAF0D1')
                              )

    active.conditional_formatting.add('G7:G125',
                ColorScaleRule(start_type='percentile', start_value=10, start_color='FFFFFF',
                            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                            end_type='percentile', end_value=95, end_color='AAF0D1')
                              )


    mintFill = PatternFill(start_color='AAF0D1',
                end_color='AAF0D1',
                fill_type='solid')
    active.conditional_formatting.add('I7:I125',
                CellIsRule(operator='greaterThan', formula=[49], stopIfTrue=True, fill=mintFill))

    active.conditional_formatting.add('K7:K125',
                ColorScaleRule(start_type='percentile', start_value=10, start_color='FFFFFF',
                            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                            end_type='percentile', end_value=95, end_color='AAF0D1')
                              )
    active.conditional_formatting.add('L7:L125',
                ColorScaleRule(start_type='percentile', start_value=10, start_color='FFFFFF',
                            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                            end_type='percentile', end_value=95, end_color='AAF0D1')
                              )
    active.conditional_formatting.add('M7:M125',
                ColorScaleRule(start_type='percentile', start_value=10, start_color='FFFFFF',
                            mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                            end_type='percentile', end_value=95, end_color='AAF0D1')
                              )

    #     dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='EE1111', end_color='EE1111'))
    #     rule = Rule(type='cellIs', dxf=dxf, formula=["10"])

    login=[]
    login_prod=[]
    login_nonprod=[]
    signers=[]
    non_signers=[]
    for (mmm,nnn) in zip(df25['Interactions'],df25['Signed']):
        if mmm > 0:
            login.append(1)
        if mmm > 0 and nnn > 0:
            login_prod.append(1)
        if mmm > 0 and nnn == 0:
            login_nonprod.append(1)
        if nnn > 0:
            signers.append(1)
        if nnn == 0:
            non_signers.append(1)
    if len(df25) > 0:
        users = float((len(login)/len(df25)))
        if len(signers) > 0:
            prod_users = float((len(login_prod)/len(signers)))
        else:
            prod_users = 0.0
        if len(non_signers) > 0:
            non_prod_users = float((len(login_nonprod)/len(non_signers)))
        else:
            if len(login_nonprod) > 0:
                non_prod_users = 1.0
            else:
                non_prod_users = 0.0
    else:
        users = 0
        if len(signers) > 0:
            prod_users = float((len(login_prod)/len(signers)))
        else:
            prod_users = 0.0
        if len(non_signers) > 0:
            non_prod_users = float((len(login_nonprod)/len(non_signers)))
        else:
            if len(login_nonprod) > 0:
                non_prod_users = 1.0
            else:
                non_prod_users = 0.0

    active['K1'] = f"Using Canvass"
    active['L1'] = users
    active['L1'].number_format = '0%'
    active['K2'] = f"Closers Using Canvass"
    active['L2'] = prod_users
    active['L2'].number_format = '0%'
    active['K3'] = f"Non-Closers Using Canvass"
    active['L3'] = non_prod_users
    active['L3'].number_format = '0%'
    active['K1'].alignment = Alignment(horizontal='right')
    active['K2'].alignment = Alignment(horizontal='right')
    active['K3'].alignment = Alignment(horizontal='right')
    active['L1'].alignment = Alignment(horizontal='center')
    active['L2'].alignment = Alignment(horizontal='center')
    active['L3'].alignment = Alignment(horizontal='center')

    rows = dataframe_to_rows(df25,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    for (aa) in list(range(7,125)):
        active[f'F{aa}'].number_format = '0%'
        active[f'H{aa}'].number_format = '0%'
        active[f'J{aa}'].number_format = '0%'

    #     office_days_to_sale=days_to_sale.loc[days_to_sale['office_id'] == i]
    #     fig = px.scatter(office_days_to_sale, title='Recruiting Efficiency',x="Days to First Sale", y="Signed Total",color='Rep Status')
    #     fig.update_layout(margin=dict(l=10, r=10, t=100, b=10),
    #                       legend=dict(orientation="h",
    #                                 yanchor="bottom",
    #                                 y=1.01,
    #                                 xanchor="left",
    #                                 x=0))
    #     fig.update_xaxes(
    #         dtick="M1",
    #         tickformat="%b%Y")

    #     fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_recruiting_efficiency.png")

    #     img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_recruiting_efficiency.png")
    #     img.width = 500
    #     img.height = 300
    #     active.add_image(img,'K4')

    #     fig = px.scatter(office_days_to_sale, title='Recruiting Efficiency over Time',x="Date Onboarded", y="Days to First Sale",color='Rep Status')
    #     fig.update_layout(margin=dict(l=10, r=10, t=100, b=10),
    #                       legend=dict(orientation="h",
    #                                 yanchor="bottom",
    #                                 y=1.01,
    #                                 xanchor="left",
    #                                 x=0))
    #     fig.update_xaxes(
    #         dtick="M1",
    #         tickformat="%b%Y")

    #     fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_recruiting_efficiency_overtime.png")

    #     img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_recruiting_efficiency_overtime.png")
    #     img.width = 500
    #     img.height = 300
    #     active.add_image(img,'K20')


    ##########################################################################################
    ##########################################################################################
    #######################################Waterfall##########################################
    ##########################################################################################
    ##########################################################################################
    sheet = wb.create_sheet("Waterfall")
    active = wb["Waterfall"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Waterfall"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    df36=df35.loc[df35['office_id'] == i]
    df36=df36[['Month (Cohort)', 'Signed', 'Approved', 'M1', 'M2']]
    rows = dataframe_to_rows(df36,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    active.cell(row=6, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=6).border = Border(bottom=Side(style='thin'))

    active['B6'].font = Font(bold=True)
    active['C6'].font = Font(bold=True)
    active['D6'].font = Font(bold=True)
    active['E6'].font = Font(bold=True)
    active['F6'].font = Font(bold=True)

    active.column_dimensions['B'].width = 15
    active.column_dimensions['C'].width = 15
    active.column_dimensions['D'].width = 17
    active.column_dimensions['E'].width = 15
    active.column_dimensions['F'].width = 15

    affectedcells1 = list(range(6, 30))
    for rownum in affectedcells1:
        active[f'C{rownum}'].alignment = Alignment(horizontal='center')
        active[f'D{rownum}'].alignment = Alignment(horizontal='center')
        active[f'E{rownum}'].alignment = Alignment(horizontal='center')
        active[f'F{rownum}'].alignment = Alignment(horizontal='center')
        if rownum >= 18:
            active[f'D{rownum}'].number_format = '0%'
            active[f'E{rownum}'].number_format = '0%'
            active[f'F{rownum}'].number_format = '0%'

    app=[]
    nunha=[]
    m1=[]
    m2=[]
    for e,r,t,y in zip(df36['Signed'],df36['Approved'],df36['M1'],df36['M2']):
        nunha.append('')
        if e != 0:
            app.append(float(r/e))
        else:
            app.append("")
        if r != 0:
            m1.append(float(t/r))
            m2.append(float(y/r))
        else:
            m1.append('')
            m2.append('')
    df36['']=nunha
    df36['Signed to Approved'] = app
    df36['Approved to M1'] = m1
    df36['Approved to M2'] = m2

    df37 = df36[['Month (Cohort)','','Signed to Approved','Approved to M1','Approved to M2']]
    rows = dataframe_to_rows(df37,index=False)
    for r_idx, row in enumerate(rows, 17):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    active['B17'].font = Font(bold=True)
    active['C17'].font = Font(bold=True)
    active['D17'].font = Font(bold=True)
    active['E17'].font = Font(bold=True)
    active['F17'].font = Font(bold=True)

    active.cell(row=17, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=17, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=17, column=4).border = Border(bottom=Side(style='thin'))
    active.cell(row=17, column=5).border = Border(bottom=Side(style='thin'))
    active.cell(row=17, column=6).border = Border(bottom=Side(style='thin'))


    ##########################################################################################
    ##########################################################################################
    ############################################Tiles#########################################
    ##########################################################################################
    ##########################################################################################
    sheet = wb.create_sheet(f"Tiles")
    active = wb[f"Tiles"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Tiles"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    df52 = df24.loc[df24['office_id'] == i]
    df52 = df52.loc[df24['Days with LGCY'] != 'Not Onboarded']

    s=[]
    for (one,two) in zip(df52['Signed'],df52['Mentored']):
        s.append(one + two)
    df52['Total'] = s
    total=sum(df52['Total'])
    if total >= 4:

        perc=[]
        for three in df52['Total']:
            perc.append(three/total)
        df52['Contribution']=perc

        threemo = df52.loc[df52['Days with LGCY'] < 90]
        sixmo = df52.loc[(df52['Days with LGCY'] >= 90) & (df52['Days with LGCY'] < 180)]
        oney = df52.loc[(df52['Days with LGCY'] >= 180) & (df52['Days with LGCY'] < 365)]
        mtone = df52.loc[df52['Days with LGCY'] >= 365]

        affectedcells1 = list(range(6, 12))
        for rownum in affectedcells1:
            active[f'C{rownum}'].alignment = Alignment(horizontal='center')
            active[f'F{rownum}'].alignment = Alignment(horizontal='center')

        active['B6'] = f"{monthname} Signed Deals by Experience"
        active.merge_cells('B6:C6')
        active['B6'].alignment = Alignment(vertical='center',horizontal='center')
        active.cell(row=6, column=2).border = Border(bottom=Side(style='thin'))
        active.cell(row=6, column=3).border = Border(bottom=Side(style='thin'))
        active['B7'].font = Font(bold=True)
        active.cell(row=7, column=2).border = Border(right=Side(style='thin'))
        active['B8'].font = Font(bold=True)
        active.cell(row=8, column=2).border = Border(right=Side(style='thin'))
        active['B9'].font = Font(bold=True)
        active.cell(row=9, column=2).border = Border(right=Side(style='thin'))
        active['B10'].font = Font(bold=True)
        active.cell(row=10, column=2).border = Border(right=Side(style='thin'))
        active.column_dimensions['B'].width = 25
        active.column_dimensions['C'].width = 20

        active['B7'] = "% From Reps 0-3 Months"
        active['B8'] = "% From Reps 3-6 Months"
        active['B9'] = "% From Reps 6-12 Months"
        active['B10'] = "% From Reps 12+ Months"
        active['C7'] = float(sum(threemo['Contribution']))
        active['C8'] = float(sum(sixmo['Contribution']))
        active['C9'] = float(sum(oney['Contribution']))
        active['C10'] = float(sum(mtone['Contribution']))
        active['C7'].number_format = '0%'
        active['C8'].number_format = '0%'
        active['C9'].number_format = '0%'
        active['C10'].number_format = '0%'

        active['E6'] = f"{monthname} Signed Deals by Rep Share"
        active.merge_cells('E6:F6')
        active['E6'].alignment = Alignment(vertical='center',horizontal='center')
        active.cell(row=6, column=5).border = Border(bottom=Side(style='thin'))
        active.cell(row=6, column=6).border = Border(bottom=Side(style='thin'))
        active['E7'].font = Font(bold=True)
        active.cell(row=7, column=5).border = Border(right=Side(style='thin'))
        active.cell(row=8, column=5).border = Border(right=Side(style='thin'))
        active.cell(row=9, column=5).border = Border(right=Side(style='thin'))
        active.cell(row=10, column=5).border = Border(right=Side(style='thin'))
        active.cell(row=11, column=5).border = Border(right=Side(style='thin'))
        active.column_dimensions['E'].width = 25
        active.column_dimensions['F'].width = 20
        active['F7'].font = Font(bold=True)
        active.cell(row=7, column=5).border = Border(bottom=Side(style='thin'))
        active.cell(row=7, column=6).border = Border(bottom=Side(style='thin'))

        df52=df52.sort_values(by=['Contribution'],ascending=False)
        total1 = 0
        count_reps = 0
        rep_breakdown = []
        rep_p_breakdown = []
        for rep_perc in df52['Contribution']:
            old_total = total1
            total1 = total1 + rep_perc
            old_count = count_reps
            count_reps = count_reps + 1
            if total1 >= .2500:
                rep_breakdown.append(old_count)
                rep_p_breakdown.append(total1 - rep_perc)
                total1 = rep_perc
                count_reps = 1

        active['E7'] = "% of Deals"
        active['E8'] = f"Top 25%"
        active['E9'] = f"Middle 25%"
        active['E10'] = f"Middle 25%"
        active['E11'] = f"Bottom 25%"

        active['F7'] = "Signed by"
        try:
            active['F8'] = f"{int(rep_breakdown[0])} Rep(s)"
            active['F9'] = f"{int(rep_breakdown[1])} Rep(s)"
            active['F10'] = f"{int(rep_breakdown[2])} Rep(s)"
            active['F11'] = f"{int(rep_breakdown[3])} Rep(s)"
        except:
            active['B6'] = f"Insufficient Data"
            active.merge_cells('F8:F11')
            active['F8'].alignment = Alignment(vertical='center',horizontal='center')



    else:
        active['B6'] = f"Insufficient Data"
        active.merge_cells('B6:C6')

    ##########################################################################################
    ##########################################################################################
    #######################################Manager Performance################################
    ##########################################################################################
    ##########################################################################################
    sheet = wb.create_sheet("Managers")
    active = wb["Managers"]
    active.sheet_view.showGridLines = False

    active['B1'] = f"{j} Managers"
    b1 = active['B1']
    b1.font = Font(size=18)
    b1.alignment = Alignment(vertical='center',horizontal='center')
    active.merge_cells('B1:H4')

    img = Image('/Users/andy/Documents/LGCY/lgcy_logo_new.png')
    img.width = 85
    img.height = 85
    active.merge_cells('A1:A4')
    active.add_image(img, 'A1')

    df44=df43.loc[df43['office_id'] == i]
    df44=df44[['rep_name','signed','installed']]
    df44=df44.sort_values(by=['signed','installed'],ascending=False)
    df44 = df44.rename(index=str, columns={'rep_name':'Manager','signed':'Signed','installed':'Installed'})
    rows = dataframe_to_rows(df44,index=False)
    for r_idx, row in enumerate(rows, 6):
        for c_idx, value in enumerate(row, 2):
            active.cell(row=r_idx, column=c_idx, value=value)

    active['B6'].font = Font(bold=True)
    active['C6'].font = Font(bold=True)
    active['D6'].font = Font(bold=True)

    active.column_dimensions['B'].width = 15

    active.cell(row=6, column=2).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=3).border = Border(bottom=Side(style='thin'))
    active.cell(row=6, column=4).border = Border(bottom=Side(style='thin'))

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df44['Manager'],
        y=df44['Signed'],
        name='Signed'
    ))
    fig.add_trace(go.Bar(
        x=df44['Manager'],
        y=df44['Installed'],
        name='Installed'
        #marker_color='lightsalmon'
    ))

    # Here we modify the tickangle of the xaxis, resulting in rotated labels.
    fig.update_layout(title=f'{monthname} Performance',barmode='group',
                        margin=dict(t=90,b=20,l=10,r=10),
                        legend=dict(orientation="h",
                                yanchor="bottom",
                                y=1.01,
                                xanchor="left",
                                x=0))#xaxis_tickangle=-45
    fig.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_manager_performance.png")

    img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_manager_performance.png")
    img.width = 800
    img.height = 400
    active.add_image(img,'H1')

    df49 = df48.loc[df48['office_id'] == i]
    if len(df49) > 0:
        fig2 = px.line(df49, x="Date", y="Signed", color='Rep')
        fig2.update_layout(title=f'{thisyear} Manager Production')

        fig2.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_manager_signed_{thisyear}.png")

        img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_manager_signed_{thisyear}.png")
        img.width = 750
        img.height = 500
        active.add_image(img,'B22')

    df51 = df50.loc[df50['office_id'] == i]
    if len(df51) > 0:
        fig3 = px.line(df51, x="Date", y="Installed", color='Rep')
    #     fig = px.pie(df7, values='Approved', names='Install Partner', color='Install Partner',
    #              color_discrete_sequence=["#0DABE3","#03254C", "#1167B1", "blue", "#187BCD", "#D0EFFF"])

        fig3.update_layout(margin=dict(t=10),
                           showlegend=False)
    #                       legend=dict(orientation="h",
    #                                 yanchor="bottom",
    #                                 y=1.01,
    #                                 xanchor="left",x=0)
                                    )
    #     fig.update_layout(legend= {'itemsizing': 'constant'})

        fig3.write_image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_manager_installed_{thisyear}.png")

        img = Image(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/images/{j}_manager_installed_{thisyear}.png")
        img.width = 650
        img.height = 400
        active.add_image(img,'L25')


    wb.save(f"/Users/andy/Dropbox (LGCY)/MBRs/{monthname}/{j}.xlsx")
